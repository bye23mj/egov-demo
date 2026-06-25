#!/usr/bin/env python3
# 선박기본정보 DA 산출물 7종 + 표준검토결과 + 산출물관리대장 + DDL 생성
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BASE = "/Users/ai/vscode/egov-demo/docs/04. db-deliverables/선박기본정보"
DDL  = "/Users/ai/vscode/egov-demo/docs/05. db-build/선박기본정보/ddl/01_tb_vsl_bsc.sql"
PFX  = "선박기본정보-"
TBL_KO, TBL_EN = "선박기본정보", "TB_VSL_BSC"

# (한글컬럼, 영문컬럼, 타입, 길이, NULL허용, 키, 코드성, 표준단어출처, 설명)
COLS = [
 ("선박기본정보ID","VSL_BSC_ID","NUMBER","",      "N","PK","","인조키","주식별자(SEQUENCE 채번)"),
 ("구분","DIV_NM","VARCHAR2","20",                 "Y","", "Y","구분(SE)","어선/일반선 대분류"),
 ("선박번호","VSL_NO","VARCHAR2","20",             "Y","IX","","선박(VSL)+번호(NO)","선박 식별번호(중복허용)"),
 ("선박명","VSL_NM","VARCHAR2","200",              "Y","", "","선박(VSL)+명","선박 명칭"),
 ("관할지사","CMPTNC_BROF_NM","VARCHAR2","60",     "Y","", "Y","관할(CMPTNC)+지사(BROF)","관할 지사명"),
 ("선박구분","VSL_SE_NM","VARCHAR2","30",          "Y","", "Y","선박(VSL)+구분(SE)","기선/부선 등 선종"),
 ("선박재질","VSL_MTPRPT_NM","VARCHAR2","30",      "Y","", "Y","선박(VSL)+재질(MTPRPT)","FRP/강 등 재질"),
 ("선적항","SHPM_PRT_NM","VARCHAR2","100",         "Y","", "","선적(SHPM)+항(PRT)","선적항"),
 ("길이","LEN","NUMBER","10,2",                    "Y","", "","길이(LEN)","선박 길이(m)"),
 ("너비","BRTH","NUMBER","10,2",                   "Y","", "","너비(BRTH)*KOMSA","선박 너비(m)"),
 ("깊이","DPTH","NUMBER","10,2",                   "Y","", "","깊이(DPTH)","선박 깊이(m)"),
 ("총톤수","TOT_TON","NUMBER","13,2",              "Y","", "","총(TOT)+톤수(TON)","총톤수(t)"),
 ("주기관종류1","MENG_KND1_NM","VARCHAR2","40",    "Y","", "Y","주기관(MENG)*+종류(KND)","주기관1 종류"),
 ("주기관1마력","MENG1_HP","NUMBER","10",          "Y","", "","주기관(MENG)*+마력(HP)*","주기관1 마력"),
 ("주기관1KW","MENG1_KW","NUMBER","10",            "Y","", "","주기관(MENG)*+KW","주기관1 출력(KW)"),
 ("주기관1_대","MENG1_CNT","NUMBER","5",           "Y","", "","주기관(MENG)*+대수(CNT)","주기관1 대수"),
 ("주기관종류2","MENG_KND2_NM","VARCHAR2","40",    "Y","", "Y","주기관(MENG)*+종류(KND)","주기관2 종류"),
 ("주기관2마력","MENG2_HP","NUMBER","10",          "Y","", "","주기관(MENG)*+마력(HP)*","주기관2 마력"),
 ("주기관2KW","MENG2_KW","NUMBER","10",            "Y","", "","주기관(MENG)*+KW","주기관2 출력(KW)"),
 ("주기관2_대","MENG2_CNT","NUMBER","5",           "Y","", "","주기관(MENG)*+대수(CNT)","주기관2 대수"),
 ("진수일","LNCH_DE","DATE","",                    "Y","", "","진수(LNCH)+일자(DE)","진수일자"),
 ("용도1","USG1_NM","VARCHAR2","60",               "Y","", "","용도(USG)","용도1"),
 ("용도2","USG2_NM","VARCHAR2","60",               "Y","", "","용도(USG)","용도2"),
 ("최종검사일","LAST_INSP_DE","DATE","",           "Y","", "","최종(LAST)+검사(INSP)+일(DE)","최종검사일자"),
 ("차기검사종류1","NXTM_INSP_KND1_NM","VARCHAR2","20","Y","","Y","차기(NXTM)+검사(INSP)+종류(KND)","차기검사1 종류"),
 ("차기검사일1","NXTM_INSP_DT1","DATE","",         "Y","", "","차기(NXTM)+검사(INSP)+일시(DT)","차기검사1 일시"),
 ("차기검사종류2","NXTM_INSP_KND2_NM","VARCHAR2","20","Y","","Y","차기(NXTM)+검사(INSP)+종류(KND)","차기검사2 종류"),
 ("차기검사일2","NXTM_INSP_DT2","DATE","",         "Y","", "","차기(NXTM)+검사(INSP)+일시(DT)","차기검사2 일시"),
 ("차기검사종류3","NXTM_INSP_KND3_NM","VARCHAR2","20","Y","","Y","차기(NXTM)+검사(INSP)+종류(KND)","차기검사3 종류"),
 ("차기검사일3","NXTM_INSP_DT3","DATE","",         "Y","", "","차기(NXTM)+검사(INSP)+일시(DT)","차기검사3 일시"),
 ("차기검사종류4","NXTM_INSP_KND4_NM","VARCHAR2","20","Y","","Y","차기(NXTM)+검사(INSP)+종류(KND)","차기검사4 종류"),
 ("차기검사일4","NXTM_INSP_DT4","DATE","",         "Y","", "","차기(NXTM)+검사(INSP)+일시(DT)","차기검사4 일시"),
 ("차기검사종류5","NXTM_INSP_KND5_NM","VARCHAR2","20","Y","","Y","차기(NXTM)+검사(INSP)+종류(KND)","차기검사5 종류"),
 ("차기검사일5","NXTM_INSP_DT5","DATE","",         "Y","", "","차기(NXTM)+검사(INSP)+일시(DT)","차기검사5 일시"),
 ("증서만료일자","CERT_EXPRY_DE","DATE","",        "Y","", "","증서(CERT)+만료(EXPRY)+일자(DE)","증서 만료일자"),
 ("계선여부","LAYUP_YN","CHAR","1",                "Y","", "Y","계선(LAYUP)*+여부(YN)","계선(운항정지) 여부"),
 ("운항중지시작일","NVGT_HLT_BGNG_DE","DATE","",   "Y","", "","운항(NVGT)+중지(HLT)+시작(BGNG)+일(DE)","운항중지 시작일"),
 ("운항중지종료일","NVGT_HLT_END_DE","DATE","",    "Y","", "","운항(NVGT)+중지(HLT)+종료(END)+일(DE)","운항중지 종료일"),
 ("여객_특등급","PSNR_SPCL_GRD","NUMBER","6",      "Y","", "","여객(PSNR)+특(SPCL)*+등급(GRD)","특등급 정원"),
 ("여객_1등급","PSNR_GRD1","NUMBER","6",           "Y","", "","여객(PSNR)+등급(GRD)","1등급 정원"),
 ("여객_2등급","PSNR_GRD2","NUMBER","6",           "Y","", "","여객(PSNR)+등급(GRD)","2등급 정원"),
 ("여객_3등급","PSNR_GRD3","NUMBER","6",           "Y","", "","여객(PSNR)+등급(GRD)","3등급 정원"),
 ("여객_소계","PSNR_STOT","NUMBER","8",            "Y","", "","여객(PSNR)+소계(STOT)","여객 소계"),
 ("선원수","CREW_CNT","NUMBER","6",                "Y","", "","선원(CREW)+수(CNT)","선원 수"),
 ("임시승선자수","TMPR_EMKT_CNT","NUMBER","6",     "Y","", "","임시(TMPR)+승선(EMKT)+수(CNT)","임시승선자 수"),
 ("최대승선인원합계","MAX_EMKT_PRSNL_SUM","NUMBER","8","Y","","","최대(MAX)+승선(EMKT)+인원(PRSNL)+합계(SUM)","최대승선인원 합계"),
 ("적재일시","LOAD_DT","TIMESTAMP","",             "N","", "","적재(LOAD)+일시(DT)","적재 추적용(DEFAULT SYSTIMESTAMP)"),
]

HDR = Font(bold=True, color="FFFFFF"); HFILL = PatternFill("solid", fgColor="305496")
WRAP = Alignment(wrap_text=True, vertical="top"); CEN = Alignment(horizontal="center", vertical="center")
thin = Side(style="thin", color="BFBFBF"); BORD = Border(thin,thin,thin,thin)

def sheet(wb, title, headers, rows, widths=None):
    ws = wb.active if wb.worksheets and wb.active.max_row==1 and wb.active.max_column==1 and wb.active['A1'].value is None and title=="_first" else wb.create_sheet()
    return ws

def make(name, headers, rows, widths, title="Sheet1", note=None):
    wb = Workbook(); ws = wb.active; ws.title = title
    r0 = 1
    if note:
        ws.cell(1,1,note).font = Font(bold=True, size=11); r0 = 2
    for c,h in enumerate(headers,1):
        cell = ws.cell(r0,c,h); cell.font=HDR; cell.fill=HFILL; cell.alignment=CEN; cell.border=BORD
    for ri,row in enumerate(rows, r0+1):
        for ci,val in enumerate(row,1):
            cell = ws.cell(ri,ci,val); cell.alignment=WRAP; cell.border=BORD
    for ci,w in enumerate(widths,1):
        ws.column_dimensions[chr(64+ci)].width = w
    ws.row_dimensions[r0].height = 22
    path = os.path.join(BASE, PFX+name+".xlsx"); wb.save(path); return path

made = []

# 1) 표준검토결과
rows = [(k, e, src, ("등록(KOMSA)" if "*" in src else "공공표준 일치"), "확정") for (k,e,t,l,n,key,cd,src,desc) in COLS if e!="VSL_BSC_ID"]
made.append(make("표준검토결과", ["한글컬럼","영문컬럼(확정)","표준단어 구성","표준원천","판정"], rows,
                 [16,22,40,16,10], "표준검토결과",
                 note="선박기본정보 표준화 검토결과 (공공표준 우선, * = KOMSA 기관표준 신규등록: 너비/마력/계선/주기관/특)"))

# 2) 논리 ERD (Mermaid)
log_attrs = "\n".join(f'    {"+ " if key=="PK" else "  "}{k}' for (k,e,t,l,n,key,cd,src,desc) in COLS)
log_mmd = f"erDiagram\n  선박기본정보 {{\n{log_attrs}\n  }}"
made.append(make("논리데이터모델다이어그램", ["구분","내용"],
                 [("표기법","Mermaid erDiagram"),("엔터티","선박기본정보 (단일, 반복그룹 비정규화-적재형)"),
                  ("주식별자","선박기본정보ID (인조키)"),("ERD",log_mmd)],
                 [14,90], "논리ERD"))

# 3) 물리 ERD (Mermaid)
def mtype(t,l): return f'"{t}({l})"' if l else f'"{t}"'
phy_attrs = "\n".join(f'    {mtype(t,l)} {e} {"PK" if key=="PK" else ("FK" if key=="FK" else "")}'.rstrip()
                      for (k,e,t,l,n,key,cd,src,desc) in COLS)
phy_mmd = f"erDiagram\n  {TBL_EN} {{\n{phy_attrs}\n  }}"
made.append(make("물리데이터모델다이어그램", ["구분","내용"],
                 [("표기법","Mermaid erDiagram"),("테이블",f"{TBL_EN} ({TBL_KO})"),
                  ("PK","PK_TB_VSL_BSC (VSL_BSC_ID)"),("INDEX","IX_TB_VSL_BSC_VSL_NO (VSL_NO, 비유일)"),
                  ("SEQUENCE","SEQ_TB_VSL_BSC"),("ERD",phy_mmd)],
                 [14,90], "물리ERD"))

# 4) 엔터티정의서
made.append(make("엔터티정의서", ["엔터티명(한글)","엔터티명(영문)","정의","주식별자","발생주기","개인정보","비고"],
                 [(TBL_KO,TBL_EN,"선박 1척의 제원·검사·정원 기본정보","선박기본정보ID(인조키)","수시","없음",
                   "선박번호 중복으로 인조키 채택, 반복그룹 비정규화(적재형)")],
                 [16,16,40,18,10,10,40], "엔터티정의서"))

# 5) 애트리뷰트정의서
attr_rows = [(TBL_KO,k,e,t+((f"({l})") if l else ""),("필수" if n=="N" else "선택"),
             (key if key else ""),(cd if cd else ""),desc) for (k,e,t,l,n,key,cd,src,desc) in COLS]
made.append(make("애트리뷰트정의서", ["엔터티","한글속성","영문속성","도메인(타입)","필수","키","코드성","정의"],
                 attr_rows, [14,16,22,16,8,6,8,34], "애트리뷰트정의서"))

# 6) 데이터베이스정의서
made.append(make("데이터베이스정의서", ["항목","값"],
                 [("DBMS","Oracle Database 11g XE"),("인스턴스/SID","XE"),("스키마(계정)","komsa"),
                  ("캐릭터셋","AL32UTF8"),("테이블스페이스","USERS (기본)"),
                  ("대상테이블","TB_VSL_BSC (선박기본정보)"),("원천","work/선박기본정보.csv (45컬럼, 68,933행)"),
                  ("적재방식","SQL*Loader (인조키 SEQUENCE, 날짜/숫자 정제 후 적재)")],
                 [18,70], "데이터베이스정의서"))

# 7) 테이블정의서
made.append(make("테이블정의서", ["한글테이블명","영문테이블명","스키마","주식별자","컬럼수","제약/인덱스","설명"],
                 [(TBL_KO,TBL_EN,"komsa","VSL_BSC_ID",str(len(COLS)),
                   "PK_TB_VSL_BSC / IX_TB_VSL_BSC_VSL_NO / SEQ_TB_VSL_BSC","선박 기본정보 단일 적재 테이블")],
                 [16,16,10,14,8,40,30], "테이블정의서"))

# 8) 컬럼정의서
col_rows = [(str(i+1),k,e,t+((f"({l})") if l else ""),("N" if n=="N" else "Y"),(key if key else ""),
            (cd if cd else ""),src,desc) for i,(k,e,t,l,n,key,cd,src,desc) in enumerate(COLS)]
made.append(make("컬럼정의서", ["순번","한글컬럼명","영문컬럼명","데이터타입","NULL","키","코드성","표준단어","설명"],
                 col_rows, [6,16,22,15,6,6,8,34,30], "컬럼정의서"))

# 9) 산출물관리대장
ledger = [
 ("1","데이터모델링","선박기본정보-db-modeling.md","적합","DA"),
 ("2","표준화검토","선박기본정보-표준검토결과.xlsx","확정(KOMSA 5건 등록)","metadata-agent"),
 ("3","데이터구조검증","선박기본정보-구조검증결과.md","적합(보완 2건 관찰)","DA"),
 ("4-1","논리ERD","선박기본정보-논리데이터모델다이어그램.xlsx","완료","DA"),
 ("4-2","물리ERD","선박기본정보-물리데이터모델다이어그램.xlsx","완료","DA"),
 ("4-3","엔터티정의서","선박기본정보-엔터티정의서.xlsx","완료","DA"),
 ("4-4","애트리뷰트정의서","선박기본정보-애트리뷰트정의서.xlsx","완료","DA"),
 ("4-5","데이터베이스정의서","선박기본정보-데이터베이스정의서.xlsx","완료","DA"),
 ("4-6","테이블정의서","선박기본정보-테이블정의서.xlsx","완료","DA"),
 ("4-7","컬럼정의서","선박기본정보-컬럼정의서.xlsx","완료","DA"),
 ("5","변경관리(ADR/마이그레이션)","adr/ADR-0001-선박기본정보-적재형설계.md / ddl/01_tb_vsl_bsc.sql","완료","DA"),
 ("6","데이터품질부적합대장","선박기본정보-데이터품질부적합대장.xlsx","적재 후 작성","DA"),
 ("7","데이터분석노트","선박기본정보-데이터분석노트.md","적재 후 작성","DA"),
]
made.append(make("산출물관리대장", ["순번","산출물구분","파일명","상태","담당"], ledger,
                 [8,24,52,24,16], "산출물관리대장"))

# DDL 생성
def ddl_type(t,l): return f"{t}({l})" if l else t
lines = ["-- 선박기본정보 테이블 DDL (자동생성: gen_vsl_deliverables.py)",
         "-- 대상: Oracle 11g XE / 스키마: komsa", "WHENEVER SQLERROR CONTINUE",
         f"DROP TABLE {TBL_EN} CASCADE CONSTRAINTS;", "DROP SEQUENCE SEQ_TB_VSL_BSC;",
         "WHENEVER SQLERROR EXIT SQL.SQLCODE", f"CREATE TABLE {TBL_EN} ("]
coldefs = []
for (k,e,t,l,n,key,cd,src,desc) in COLS:
    d = f"  {e} {ddl_type(t,l)}"
    if e=="LOAD_DT": d += " DEFAULT SYSTIMESTAMP"
    if n=="N": d += " NOT NULL"
    coldefs.append(d)
coldefs.append("  CONSTRAINT PK_TB_VSL_BSC PRIMARY KEY (VSL_BSC_ID)")
lines.append(",\n".join(coldefs)); lines.append(");")
lines.append("CREATE SEQUENCE SEQ_TB_VSL_BSC START WITH 1 INCREMENT BY 1 NOCACHE;")
lines.append(f"CREATE INDEX IX_TB_VSL_BSC_VSL_NO ON {TBL_EN} (VSL_NO);")
# 코멘트
lines.append(f"COMMENT ON TABLE {TBL_EN} IS '선박기본정보';")
for (k,e,t,l,n,key,cd,src,desc) in COLS:
    lines.append(f"COMMENT ON COLUMN {TBL_EN}.{e} IS '{k}';")
lines.append("EXIT")
os.makedirs(os.path.dirname(DDL), exist_ok=True)
open(DDL,"w",encoding="utf-8").write("\n".join(lines)+"\n")

print("생성 산출물:")
for p in made: print("  ", os.path.basename(p))
print("DDL:", DDL)
print("컬럼수:", len(COLS))
