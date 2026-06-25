#!/usr/bin/env python3
# 선박기본정보 정규화 모델 산출물 7종 + DDL 생성 (부모1 + 자식3)
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BASE = "/Users/ai/vscode/egov-demo/docs/04. db-deliverables/선박기본정보/정규화"
DDL  = "/Users/ai/vscode/egov-demo/docs/05. db-build/선박기본정보/ddl/03_normalize_split.sql"
PFX  = "선박기본정보-정규화-"

# 자식 테이블 정의: (영문, 한글, [(컬럼, 타입, 키, 표준, 설명)...], 시퀀스, 복합UK)
PARENT_KEEP = "VSL_BSC_ID, 구분, 선박번호, 선박명, 관할지사, 선박구분, 선박재질, 선적항, 길이, 너비, 깊이, 총톤수, 진수일, 용도1·2, 최종검사일, 증서만료일자, 계선여부, 운항중지시작/종료일, 여객소계, 선원수, 임시승선자수, 최대승선인원합계, 적재일시"
PARENT_DROP = "주기관종류1/마력1/KW1/대1, 주기관종류2/마력2/KW2/대2, 차기검사종류1~5, 차기검사일1~5, 여객_특/1/2/3등급"

CHILDREN = [
 ("TB_VSL_MENG","선박주기관","SEQ_TB_VSL_MENG","(VSL_BSC_ID, MENG_SEQ)",[
   ("VSL_MENG_ID","NUMBER","","PK","인조키","주식별자(SEQ)"),
   ("VSL_BSC_ID","NUMBER","","FK","-","→TB_VSL_BSC"),
   ("MENG_SEQ","NUMBER","2","UK","주기관(MENG)+순서(SEQ)","주기관 순번 1~2"),
   ("MENG_KND_NM","VARCHAR2","40","","주기관+종류(KND)","주기관 종류"),
   ("MENG_HP","NUMBER","10","","주기관+마력(HP)","마력"),
   ("MENG_KW","NUMBER","10","","주기관+KW","출력(KW)"),
   ("MENG_CNT","NUMBER","5","","주기관+수(CNT)","대수"),
 ]),
 ("TB_VSL_NXTM_INSP","선박차기검사","SEQ_TB_VSL_NXTM_INSP","(VSL_BSC_ID, NXTM_INSP_SEQ)",[
   ("VSL_NXTM_INSP_ID","NUMBER","","PK","인조키","주식별자(SEQ)"),
   ("VSL_BSC_ID","NUMBER","","FK","-","→TB_VSL_BSC"),
   ("NXTM_INSP_SEQ","NUMBER","2","UK","차기(NXTM)+검사(INSP)+순서(SEQ)","검사 순번 1~5"),
   ("NXTM_INSP_KND_NM","VARCHAR2","20","","차기검사+종류(KND)","차기검사 종류"),
   ("NXTM_INSP_DT","DATE","","","차기검사+일시(DT)","차기검사 일시"),
 ]),
 ("TB_VSL_PSNR_CPCT","선박여객정원","SEQ_TB_VSL_PSNR_CPCT","(VSL_BSC_ID, PSNR_GRD_SE)",[
   ("VSL_PSNR_CPCT_ID","NUMBER","","PK","인조키","주식별자(SEQ)"),
   ("VSL_BSC_ID","NUMBER","","FK","-","→TB_VSL_BSC"),
   ("PSNR_GRD_SE","VARCHAR2","10","UK","여객(PSNR)+등급(GRD)+구분(SE)","등급 특/1/2/3"),
   ("PSNR_CPCT_CNT","NUMBER","6","","여객+정원(PSCP)+수(CNT)","등급별 정원"),
 ]),
]

HDR=Font(bold=True,color="FFFFFF");HF=PatternFill("solid",fgColor="305496")
CEN=Alignment(horizontal="center",vertical="center");WRAP=Alignment(wrap_text=True,vertical="top")
t=Side(style="thin",color="BFBFBF");B=Border(t,t,t,t)

def make(name,headers,rows,widths,title="Sheet1",note=None):
    wb=Workbook();ws=wb.active;ws.title=title;r0=1
    if note: ws.cell(1,1,note).font=Font(bold=True,size=11);r0=2
    for c,h in enumerate(headers,1):
        x=ws.cell(r0,c,h);x.font=HDR;x.fill=HF;x.alignment=CEN;x.border=B
    for ri,row in enumerate(rows,r0+1):
        for ci,v in enumerate(row,1):
            x=ws.cell(ri,ci,v);x.alignment=WRAP;x.border=B
    for ci,w in enumerate(widths,1): ws.column_dimensions[chr(64+ci)].width=w
    p=os.path.join(BASE,PFX+name+".xlsx");wb.save(p);return p

def typ(t,l): return f"{t}({l})" if l else t
made=[]

# 1 논리 ERD (Mermaid, 관계 포함)
log=("erDiagram\n"
 "  선박기본정보 ||--o{ 선박주기관 : 보유\n"
 "  선박기본정보 ||--o{ 선박차기검사 : 보유\n"
 "  선박기본정보 ||--o{ 선박여객정원 : 보유\n"
 "  선박기본정보 { string 선박번호 string 선박명 }\n"
 "  선박주기관 { number 주기관순번 string 주기관종류 }\n"
 "  선박차기검사 { number 차기검사순번 date 차기검사일시 }\n"
 "  선박여객정원 { string 여객등급구분 number 여객정원수 }")
made.append(make("논리데이터모델다이어그램",["구분","내용"],
  [("표기법","Mermaid erDiagram"),("관계","선박기본정보(1) ──< 자식3(N)"),("ERD",log)],[14,90],"논리ERD"))

# 2 물리 ERD (Mermaid, FK 관계)
phy=["erDiagram",
 "  TB_VSL_BSC ||--o{ TB_VSL_MENG : FK_VSL_BSC_ID",
 "  TB_VSL_BSC ||--o{ TB_VSL_NXTM_INSP : FK_VSL_BSC_ID",
 "  TB_VSL_BSC ||--o{ TB_VSL_PSNR_CPCT : FK_VSL_BSC_ID"]
for en,kn,seq,uk,cols in CHILDREN:
    phy.append(f"  {en} {{")
    for c,ty,l,key,std,desc in cols:
        phy.append(f'    "{typ(ty,l)}" {c} {key if key in ("PK","FK") else ""}'.rstrip())
    phy.append("  }")
made.append(make("물리데이터모델다이어그램",["구분","내용"],
  [("표기법","Mermaid erDiagram"),("부모","TB_VSL_BSC (반복그룹 제거)"),("ERD","\n".join(phy))],[14,90],"물리ERD"))

# 3 엔터티정의서
ent=[("선박기본정보","TB_VSL_BSC","선박 제원·요약(반복그룹 제거)","VSL_BSC_ID","부모"),
     ("선박주기관","TB_VSL_MENG","선박별 주기관 N건","VSL_MENG_ID","자식 1:N"),
     ("선박차기검사","TB_VSL_NXTM_INSP","선박별 차기검사 N건","VSL_NXTM_INSP_ID","자식 1:N"),
     ("선박여객정원","TB_VSL_PSNR_CPCT","선박별 여객등급 정원 N건","VSL_PSNR_CPCT_ID","자식 1:N")]
made.append(make("엔터티정의서",["엔터티","영문테이블","정의","주식별자","관계"],ent,[16,20,40,18,12],"엔터티정의서"))

# 4 애트리뷰트정의서 (자식3)
attr=[]
for en,kn,seq,uk,cols in CHILDREN:
    for c,ty,l,key,std,desc in cols:
        attr.append((kn,c,typ(ty,l),key,std,desc))
made.append(make("애트리뷰트정의서",["엔터티","영문속성","도메인(타입)","키","표준단어","정의"],attr,[16,22,16,6,30,30],"애트리뷰트정의서"))

# 5 데이터베이스정의서
made.append(make("데이터베이스정의서",["항목","값"],
  [("DBMS","Oracle 11g XE"),("스키마","komsa"),("모델","정규화(부모1+자식3, 1:N)"),
   ("부모","TB_VSL_BSC (반복그룹 컬럼 제거)"),
   ("자식","TB_VSL_MENG / TB_VSL_NXTM_INSP / TB_VSL_PSNR_CPCT"),
   ("채번","SEQ_TB_VSL_MENG / SEQ_TB_VSL_NXTM_INSP / SEQ_TB_VSL_PSNR_CPCT"),
   ("관계","FK VSL_BSC_ID → TB_VSL_BSC (RESTRICT)")],[18,70],"데이터베이스정의서"))

# 6 테이블정의서
tbl=[("선박기본정보","TB_VSL_BSC","부모","VSL_BSC_ID","반복그룹 제거 후 요약"),]
for en,kn,seq,uk,cols in CHILDREN:
    tbl.append((kn,en,"자식(1:N)",cols[0][0],f"FK VSL_BSC_ID, UK {uk}"))
made.append(make("테이블정의서",["한글테이블","영문테이블","유형","주식별자","비고"],tbl,[16,20,12,18,34],"테이블정의서"))

# 7 컬럼정의서 (자식3)
col=[]
for en,kn,seq,uk,cols in CHILDREN:
    for i,(c,ty,l,key,std,desc) in enumerate(cols,1):
        col.append((en,str(i),c,typ(ty,l),key,std,desc))
made.append(make("컬럼정의서",["테이블","순번","영문컬럼","데이터타입","키","표준단어","설명"],col,[18,6,22,15,6,30,28],"컬럼정의서"))

# 산출물관리대장
ledger=[("1","정규화모델링","선박기본정보-정규화-db-modeling.md","적합"),
 ("2","표준화검토","정원(PSCP) 등록·순서(SEQ) 재사용","확정"),
 ("3","구조검증","선박기본정보-정규화-구조검증결과.md","적합"),
 ("4-1~7","산출물 7종","논리/물리ERD·엔터티·애트리뷰트·DB·테이블·컬럼정의서","완료"),
 ("5","변경관리","adr/ADR-0002 + ddl/03_normalize_split.sql","완료")]
made.append(make("산출물관리대장",["순번","구분","산출물","상태"],ledger,[10,18,52,16],"산출물관리대장"))

# DDL (자식 생성 + 데이터 분해 INSERT...SELECT)
L=["-- 선박기본정보 정규화 분해 DDL (자동생성)","-- 대상: Oracle 11g / komsa","-- 부모 TB_VSL_BSC 유지, 자식3 생성 후 반복그룹 데이터 분해 적재","WHENEVER SQLERROR CONTINUE"]
for en,kn,seq,uk,cols in CHILDREN:
    L.append(f"DROP TABLE {en} CASCADE CONSTRAINTS;")
    L.append(f"DROP SEQUENCE {seq};")
L.append("WHENEVER SQLERROR EXIT SQL.SQLCODE")
# 테이블 생성
for en,kn,seq,uk,cols in CHILDREN:
    defs=[]
    for c,ty,l,key,std,desc in cols:
        d=f"  {c} {typ(ty,l)}"
        if key in ("PK","FK") or c.endswith("_SEQ") or c=="PSNR_GRD_SE": d+=" NOT NULL"
        defs.append(d)
    pk=cols[0][0]
    defs.append(f"  CONSTRAINT PK_{en} PRIMARY KEY ({pk})")
    defs.append(f"  CONSTRAINT FK_{en}_BSC FOREIGN KEY (VSL_BSC_ID) REFERENCES TB_VSL_BSC(VSL_BSC_ID)")
    defs.append(f"  CONSTRAINT UK_{en} UNIQUE {uk}")
    L.append(f"CREATE TABLE {en} (\n"+",\n".join(defs)+"\n);")
    L.append(f"CREATE SEQUENCE {seq} START WITH 1 INCREMENT BY 1 NOCACHE;")
    L.append(f"CREATE INDEX IX_{en}_BSC ON {en}(VSL_BSC_ID);")
    L.append(f"COMMENT ON TABLE {en} IS '{kn}';")
# 데이터 분해 INSERT...SELECT
L.append("\n-- ▼ 데이터 분해 적재 (원천 TB_VSL_BSC 반복그룹 → 자식)")
for n in (1,2):
    L.append(f"INSERT INTO TB_VSL_MENG (VSL_MENG_ID, VSL_BSC_ID, MENG_SEQ, MENG_KND_NM, MENG_HP, MENG_KW, MENG_CNT)\n"
             f"SELECT SEQ_TB_VSL_MENG.NEXTVAL, VSL_BSC_ID, {n}, MENG_KND{n}_NM, MENG{n}_HP, MENG{n}_KW, MENG{n}_CNT\n"
             f"FROM TB_VSL_BSC WHERE MENG_KND{n}_NM IS NOT NULL OR MENG{n}_HP IS NOT NULL;")
for n in (1,2,3,4,5):
    L.append(f"INSERT INTO TB_VSL_NXTM_INSP (VSL_NXTM_INSP_ID, VSL_BSC_ID, NXTM_INSP_SEQ, NXTM_INSP_KND_NM, NXTM_INSP_DT)\n"
             f"SELECT SEQ_TB_VSL_NXTM_INSP.NEXTVAL, VSL_BSC_ID, {n}, NXTM_INSP_KND{n}_NM, NXTM_INSP_DT{n}\n"
             f"FROM TB_VSL_BSC WHERE NXTM_INSP_KND{n}_NM IS NOT NULL OR NXTM_INSP_DT{n} IS NOT NULL;")
for grd,coln in [("특","PSNR_SPCL_GRD"),("1","PSNR_GRD1"),("2","PSNR_GRD2"),("3","PSNR_GRD3")]:
    L.append(f"INSERT INTO TB_VSL_PSNR_CPCT (VSL_PSNR_CPCT_ID, VSL_BSC_ID, PSNR_GRD_SE, PSNR_CPCT_CNT)\n"
             f"SELECT SEQ_TB_VSL_PSNR_CPCT.NEXTVAL, VSL_BSC_ID, '{grd}', {coln}\n"
             f"FROM TB_VSL_BSC WHERE {coln} IS NOT NULL AND {coln} > 0;")
L.append("COMMIT;")
L.append("\n-- (선택) 부모에서 반복그룹 컬럼 제거 — 검증 후 수동 실행 권장")
L.append("-- ALTER TABLE TB_VSL_BSC DROP (MENG_KND1_NM, MENG1_HP, MENG1_KW, MENG1_CNT, MENG_KND2_NM, MENG2_HP, MENG2_KW, MENG2_CNT,")
L.append("--   NXTM_INSP_KND1_NM, NXTM_INSP_DT1, NXTM_INSP_KND2_NM, NXTM_INSP_DT2, NXTM_INSP_KND3_NM, NXTM_INSP_DT3,")
L.append("--   NXTM_INSP_KND4_NM, NXTM_INSP_DT4, NXTM_INSP_KND5_NM, NXTM_INSP_DT5, PSNR_SPCL_GRD, PSNR_GRD1, PSNR_GRD2, PSNR_GRD3);")
L.append("EXIT")
os.makedirs(os.path.dirname(DDL),exist_ok=True)
open(DDL,"w",encoding="utf-8").write("\n".join(L)+"\n")

print("생성 산출물:")
for p in made: print("  ",os.path.basename(p))
print("DDL:",os.path.basename(DDL))
