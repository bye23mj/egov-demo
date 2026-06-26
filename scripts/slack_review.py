#!/usr/bin/env python3
"""Slack 첨부 검토 모듈 — slack_bot_receiver 의 '연결 지점'에서 호출.

첨부 내용 추출 → 명령/에이전트별 규칙 검토 → Markdown 결과 + Slack 요약 생성.
전문 에이전트(da/dba/metadata)의 핵심 점검을 규칙 기반으로 수행한다(Python 단독 동작).
"""
import os
import re
import zipfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
STD_WORD_TSV = ROOT / "docs" / "03. metadata" / "공공" / "공공표준단어.tsv"
KOMSA_WORD_TSV = ROOT / "docs" / "03. metadata" / "기관" / "KOMSA표준단어.tsv"


# ── 표준단어 영문약어 사전 로드(컬럼명 표준화 매칭용) ──
def load_std_abbr():
    abbr = set()
    for tsv in (STD_WORD_TSV, KOMSA_WORD_TSV):
        if tsv.exists():
            for i, line in enumerate(tsv.read_text(encoding="utf-8").splitlines()):
                if i == 0:
                    continue
                cols = line.split("\t")
                if len(cols) > 2 and cols[2].strip():
                    abbr.add(cols[2].strip().upper())
    return abbr


# ── 첨부 내용 추출 ──
def extract(path: Path):
    """{type, tables:[{name,id,desc,columns:[{id,name,type,len,null,key}]}], text, note}."""
    ext = path.suffix.lower()
    if ext == ".xls":
        return _extract_xls(path)
    if ext == ".xlsx":
        return _extract_xlsx(path)
    if ext in (".csv", ".tsv", ".md", ".txt", ".sql"):
        return {"type": ext, "tables": [], "text": path.read_text(encoding="utf-8", errors="replace")[:20000], "note": ""}
    if ext == ".docx":
        return {"type": "docx", "tables": [], "text": _docx_text(path)[:20000], "note": ""}
    if ext == ".pdf":
        return {"type": "pdf", "tables": [], "text": "", "note": "PDF 본문 추출 생략(검토는 메타·파일명 기준)"}
    return {"type": ext, "tables": [], "text": "", "note": "지원하지 않는 형식 — 파일명만 검토"}


def _rows_to_tables(sheet_rows):
    """테이블정의서 시트 행 패턴에서 테이블 메타+컬럼을 파싱."""
    tid = tname = tdesc = ""
    cols, header_seen = [], False
    for row in sheet_rows:
        cells = [str(x).strip() for x in row]
        joined = " ".join(cells)
        if "테이블ID" in joined:
            for j, c in enumerate(cells):
                if c == "테이블ID":
                    tid = next((cells[k] for k in range(j + 1, len(cells)) if cells[k]), "")
                if c == "테이블명":
                    tname = next((cells[k] for k in range(j + 1, len(cells)) if cells[k]), "")
        elif "테이블설명" in joined and not tdesc:
            tdesc = next((c for c in cells if c and c != "테이블설명"), "")
        elif cells[:1] == ["No."] or (len(cells) > 6 and "컬럼ID" in joined and "타입" in joined):
            header_seen = True
        elif header_seen and len(cells) >= 7 and cells[1] and re.match(r"^[A-Za-z_]", cells[1]):
            cols.append({"id": cells[1], "name": cells[2], "type": cells[3],
                         "len": cells[4], "null": cells[5], "key": cells[6]})
    return {"id": tid, "name": tname, "desc": tdesc, "columns": cols}


def _extract_xls(path):
    import xlrd
    wb = xlrd.open_workbook(path)
    tables = []
    for sn in wb.sheet_names():
        sh = wb.sheet_by_name(sn)
        rows = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
        t = _rows_to_tables(rows)
        if t["columns"]:
            t["sheet"] = sn
            tables.append(t)
    return {"type": "xls", "tables": tables, "text": "", "note": ""}


def _extract_xlsx(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    tables = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = [[c if c is not None else "" for c in row] for row in ws.iter_rows(values_only=True)]
        t = _rows_to_tables(rows)
        if t["columns"]:
            t["sheet"] = sn
            tables.append(t)
    return {"type": "xlsx", "tables": tables, "text": "", "note": ""}


def _docx_text(path):
    try:
        with zipfile.ZipFile(path) as z:
            xml = z.read("word/document.xml").decode("utf-8", "replace")
        return re.sub(r"<[^>]+>", " ", xml)
    except Exception:
        return ""


# ── 검토(테이블정의서 대상: da/dba 핵심 점검) ──
def review_table_def(data, std_abbr):
    findings, tables = [], data["tables"]
    forbidden = {"DATA", "INFO", "MASTER", "TEMP", "ETC", "TEST", "OLD", "NEW"}
    total_cols = pk_missing = type_missing = std_hit = std_tot = 0
    for t in tables:
        cols = t["columns"]; total_cols += len(cols)
        has_pk = any(c["key"].upper() == "PK" for c in cols)
        if not has_pk:
            pk_missing += 1
            findings.append((t.get("sheet", t["id"]), "식별자", "보완", "PK 미정의"))
        for c in cols:
            if not c["type"]:
                type_missing += 1
            for tok in re.split(r"[_\s]+", c["id"]):
                if tok:
                    std_tot += 1
                    if tok.upper() in std_abbr:
                        std_hit += 1
        tid = (t["id"] or "").upper()
        if tid and not tid.startswith("TB"):
            findings.append((t.get("sheet", t["id"]), "테이블명", "관찰", f"테이블ID '{t['id']}' 권장 접두(TB_) 미사용"))
        if any(f in tid for f in forbidden):
            findings.append((t.get("sheet", t["id"]), "테이블명", "보완", "금지명 포함"))
    std_ratio = round(std_hit / std_tot * 100, 1) if std_tot else 0
    if type_missing:
        findings.append(("(전체)", "도메인", "보완", f"타입 미기재 컬럼 {type_missing}개"))
    verdict = "적합" if not pk_missing and type_missing == 0 else f"보완 {len(findings)}건"
    summary = {
        "tables": len(tables), "columns": total_cols, "pk_missing": pk_missing,
        "type_missing": type_missing, "std_ratio": std_ratio, "verdict": verdict,
    }
    return summary, findings


def run_review(cmd, agent, files, workdir: Path):
    """파일들을 추출·검토하고 (md_path, slack_summary) 반환."""
    workdir = Path(workdir).resolve()
    std_abbr = load_std_abbr()
    md = [f"# {cmd} 검토 결과\n", f"- 라우팅: **{agent or '미매칭'}**", f"- 첨부: {', '.join(f.name for f in files) or '없음'}\n"]
    slack_lines = [f":mag: *{cmd}* 검토 완료" + (f" — {agent}" if agent else "")]
    overall = []

    for f in files:
        data = extract(f)
        md.append(f"## {f.name} ({data['type']})")
        if data["tables"]:
            summary, findings = review_table_def(data, std_abbr)
            overall.append(summary)
            md.append(f"- 테이블 {summary['tables']} / 컬럼 {summary['columns']} / "
                      f"PK누락 {summary['pk_missing']} / 타입누락 {summary['type_missing']} / "
                      f"표준단어 매칭율 {summary['std_ratio']}%")
            md.append(f"- 판정: **{summary['verdict']}**\n")
            if findings:
                md.append("| 대상 | 범주 | 판정 | 지적 |\n|---|---|---|---|")
                for s in findings[:40]:
                    md.append(f"| {s[0]} | {s[1]} | {s[2]} | {s[3]} |")
                md.append("")
        elif data["text"]:
            md.append(f"- 텍스트 {len(data['text'])}자 추출. (구조화 검토 대상 아님 — 내용 요약 검토)\n")
        else:
            md.append(f"- {data['note'] or '본문 추출 불가'}\n")

    if overall:
        tt = sum(o["tables"] for o in overall); tc = sum(o["columns"] for o in overall)
        pk = sum(o["pk_missing"] for o in overall); tm = sum(o["type_missing"] for o in overall)
        sr = round(sum(o["std_ratio"] for o in overall) / len(overall), 1)
        slack_lines.append(f"• 테이블 {tt} / 컬럼 {tc} / PK누락 {pk} / 타입누락 {tm} / 표준매칭 {sr}%")
        issues = []
        if pk: issues.append(f"PK 미정의 {pk}개 테이블")
        if tm: issues.append(f"타입 미기재 {tm}개 컬럼")
        if sr < 70: issues.append(f"표준단어 매칭율 낮음({sr}%)")
        if issues:
            slack_lines.append("• 핵심 지적: " + " / ".join(issues[:5]))
        slack_lines.append(":arrow_right: 다음: 지적 보완 후 dba-agent로 DB 구축 / 산출물 Confluence 등록")
    else:
        slack_lines.append("• 구조화 검토 대상(테이블정의서 등) 아님 — 내용 기반 검토 필요")
        slack_lines.append(":arrow_right: 다음: 적합한 명령/첨부로 재요청")

    workdir.mkdir(parents=True, exist_ok=True)
    md_path = workdir / f"{cmd}-검토결과.md"
    md_path.write_text("\n".join(md), encoding="utf-8")
    try:
        disp = md_path.relative_to(ROOT)
    except ValueError:
        disp = md_path
    slack_lines.append(f"(상세: `{disp}`)")
    return md_path, "\n".join(slack_lines)
